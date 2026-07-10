#!/usr/bin/env python3
"""Build the "changed units only" xlsx from out/changed-units.json (written by
prices-sync.js). One row per changed date-range, prices in USD (old -> new + change).
Attached to the change email by send-alert.js. No file is written when nothing changed.
"""
import json
import os

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SRC = "out/changed-units.json"
OUT = "out/brassbell-changes.xlsx"

def main():
    if not os.path.exists(SRC):
        print("build-changes: no changed-units.json — nothing to build.")
        return
    data = json.load(open(SRC))
    if not data or not data.get("units"):
        print("build-changes: no changes — not writing a sheet.")
        if os.path.exists(OUT):
            os.remove(OUT)
        return

    date_str = data.get("dateStr", "")
    cols = [
        ("Unit ID", 12), ("Unit", 26), ("Area", 16),
        ("Start Date", 13), ("End Date", 13),
        ("Old (USD)", 11), ("New (USD)", 11), ("Change (USD)", 12),
    ]
    wb = Workbook()
    ws = wb.active
    ws.title = "Price Changes"
    ws.append([c[0] for c in cols])
    for i, (h, w) in enumerate(cols, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    rows = []
    for u in data["units"]:
        for r in u.get("ranges", []):
            old, new = r.get("oldUsd"), r.get("newUsd")
            change = (new - old) if isinstance(old, (int, float)) and isinstance(new, (int, float)) else None
            rows.append([u.get("wp"), u.get("title") or u.get("code") or u.get("wp"),
                         u.get("area") or "", r.get("from"), r.get("to"), old, new, change])
    rows.sort(key=lambda x: (str(x[0]), str(x[3])))
    for row in rows:
        ws.append(row)

    # styling
    hdr_fill = PatternFill("solid", fgColor="FF1F3B57")
    hdr_font = Font(name="Arial", bold=True, color="FFFFFFFF", size=10)
    base = Font(name="Arial", size=10)
    thin = Side(style="thin", color="FFD9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c in range(1, len(cols) + 1):
        cell = ws.cell(1, c)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[1].height = 26
    for ri in range(2, ws.max_row + 1):
        for ci in range(1, len(cols) + 1):
            cell = ws.cell(ri, ci)
            cell.font = base
            cell.border = border
            if ci in (6, 7, 8) and isinstance(cell.value, (int, float)):
                cell.number_format = "$#,##0;-$#,##0"
        chg = ws.cell(ri, 8)
        if isinstance(chg.value, (int, float)):
            chg.font = Font(name="Arial", size=10, bold=True,
                            color="FFB00020" if chg.value < 0 else "FF1B7A3D")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{chr(64 + len(cols))}{ws.max_row}"

    wb.save(OUT)
    print(f"build-changes: wrote {OUT} ({len(rows)} changed date-ranges across {len(data['units'])} units, {date_str}).")

if __name__ == "__main__":
    main()
