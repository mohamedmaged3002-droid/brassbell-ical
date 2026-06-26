#!/usr/bin/env python3
# Build the Brassbell OTA workbook from out/sheet-data.json (written by prices-sync.js)
# + data/drive-links.json (static photo-folder links). Tabs: Units, Monthly Prices, Price Ranges.
# Output path is printed on the last line so the workflow can rclone-upload it.
import json, os
from datetime import date, timedelta
from calendar import month_abbr
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
DATA = json.load(open(os.path.join(HERE, 'out', 'sheet-data.json')))
DRIVE = {int(k): v for k, v in json.load(open(os.path.join(HERE, 'data', 'drive-links.json'))).items()}
OUT = os.path.join(HERE, 'Brassbell Onboarding OTAs - Full (iCal, Photos, Price).xlsx')

FX = DATA['fx']
AS_OF = date(*map(int, DATA['dateStr'].split('-')))
WEBSITE_FX = FX
DAYS = 365
ICAL_BASE = 'https://mohamedmaged3002-droid.github.io/brassbell-ical'
SITE_BASE = 'https://www.bluekeys.co/listings'

units = DATA['units']
prices = {u['wp']: {d: int(p) for d, p in (u.get('prices') or {}).items()} for u in units}
blocked = {u['wp']: set(u.get('blocked') or []) for u in units}
meta = {u['wp']: u for u in units}

ARIAL = 'Arial'
H = Font(name=ARIAL, bold=True, color='FFFFFF', size=10)
HEAD_FILL = PatternFill('solid', fgColor='1F4E78')
NOTE = Font(name=ARIAL, italic=True, size=9, color='595959')
BODY = Font(name=ARIAL, size=10)
LINK = Font(name=ARIAL, size=9, color='1155CC', underline='single')
BLK_FONT = Font(name=ARIAL, size=10, italic=True, color='909090')
thin = Side(style='thin', color='E2E2E2')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
CEN = Alignment(horizontal='center', vertical='center')
LEFT = Alignment(horizontal='left', vertical='center')
USD_FMT = '$#,##0'
DATE_FMT = 'ddd dd-mmm-yy'
BAND = PatternFill('solid', fgColor='F2F7FB')
GREY = PatternFill('solid', fgColor='EDEDED')
NEUTRAL = PatternFill('solid', fgColor='EAF1F8')

_fc = {}
def heat_fill(t):
    t = max(0.0, min(1.0, t))
    if t <= 0.5: f = t/0.5; a = (99,190,123); b = (255,235,132)
    else: f = (t-0.5)/0.5; a = (255,235,132); b = (248,105,107)
    hexc = '%02X%02X%02X' % tuple(round(a[i]+(b[i]-a[i])*f) for i in range(3))
    if hexc not in _fc: _fc[hexc] = PatternFill('solid', fgColor=hexc)
    return _fc[hexc]

# 13 months from AS_OF
months = []
yy, mm = AS_OF.year, AS_OF.month
for _ in range(13):
    months.append((yy, mm)); mm += 1
    if mm > 12: mm = 1; yy += 1
def mlabel(y, mo): return f"{month_abbr[mo]} '{y % 100:02d}"

def iso2d(s): y, mo, da = map(int, s.split('-')); return date(y, mo, da)

def month_cell(wp, y, mo):
    pr = prices.get(wp, {})
    items = sorted((d, p) for d, p in pr.items() if int(d[:4]) == y and int(d[5:7]) == mo)
    if not items:
        blk = any(int(d[:4]) == y and int(d[5:7]) == mo for d in blocked.get(wp, set()))
        return ('blkd' if blk else '', None, False)
    distinct = {p for _, p in items}
    rep = max(distinct)
    if len(distinct) == 1:
        return (f'${rep:,}', rep, False)
    runs = []
    for d, p in items:
        day = int(d[8:10])
        if runs and runs[-1][2] == p: runs[-1][1] = day
        else: runs.append([day, day, p])
    parts = [(str(s) if s == e else f'{s}–{e}') + f': ${p:,}' for s, e, p in runs]
    return ('\n'.join(parts), rep, True)

def segments(wp):
    pr = prices.get(wp, {}); blk = blocked.get(wp, set())
    alld = sorted(set(pr) | blk)
    if not alld: return []
    start, end = iso2d(alld[0]), iso2d(alld[-1])
    out = []; cur = None; day = start
    while day <= end:
        s = day.isoformat()
        stt = ('P', pr[s]) if s in pr else (('B', None) if s in blk else ('X', None))
        if cur and cur[0] == stt: cur[2] = day
        else:
            if cur: out.append(cur)
            cur = [stt, day, day]
        day += timedelta(days=1)
    if cur: out.append(cur)
    return out

def site_url(u): return f"{SITE_BASE}/{u['slug']}" if u.get('slug') else None
def ical_url(u): return f"{ICAL_BASE}/{u['wp']}.ics"

units_sorted = sorted(units, key=lambda u: (u.get('area') or '', u['wp']))
priced_units = [u for u in units_sorted if prices.get(u['wp'])]
wb = Workbook()

# ---- Sheet 1: Units ----
ws = wb.active; ws.title = 'Units'
ws['A1'] = 'BlueKeys — Brassbell OTA Onboarding (USD)'; ws['A1'].font = Font(name=ARIAL, bold=True, size=16, color='1F4E78')
ws['A2'] = f"USD for OTA listings. Auto-refreshed daily from brassbell.net (as of {DATA['dateStr']}). See 'Monthly Prices' / 'Price Ranges'."; ws['A2'].font = NOTE
ws['A3'] = f"bluekeys.co lists these in EGP at {WEBSITE_FX} EGP/USD; OTA sheet is USD. BlueKeys adds no service/cleaning fee to Brassbell — nightly = guest total."; ws['A3'].font = NOTE
cols = ['wp_post_id','Code/Slug','Title','Area','City','Compound','Beds','Baths','Guests','Status',
        'Min nights','Nights priced','Nights blocked','Website (bluekeys.co)','iCal feed (for OTAs)','Drive photos','brassbell.net URL']
hrow = 5
for i, c in enumerate(cols, 1): ws.cell(row=hrow, column=i, value=c)
for c in range(1, len(cols)+1):
    cell = ws.cell(row=hrow, column=c); cell.font = H; cell.fill = HEAD_FILL; cell.alignment = CEN; cell.border = BORDER
ws.freeze_panes = f'A{hrow+1}'
def put_link(cell, url):
    if url: cell.value = url; cell.hyperlink = url; cell.font = LINK; cell.alignment = LEFT
r = hrow + 1
for u in units_sorted:
    wp = u['wp']; pr = prices.get(wp, {}); nb = len(blocked.get(wp, set()))
    base = [wp, u.get('slug'), u.get('title'), u.get('area'), u.get('city'), u.get('compound'),
            u.get('beds'), u.get('baths'), u.get('guests'), 'published', u.get('min_nights'), len(pr), nb]
    for i, val in enumerate(base, 1):
        cell = ws.cell(row=r, column=i, value=val); cell.font = BODY; cell.border = BORDER
        cell.alignment = LEFT if i in (2,3,4,5,6,10) else CEN
    put_link(ws.cell(row=r, column=14), site_url(u))
    put_link(ws.cell(row=r, column=15), ical_url(u))
    put_link(ws.cell(row=r, column=16), DRIVE.get(wp))
    c = ws.cell(row=r, column=17, value=u.get('source_url'))
    if u.get('source_url'): c.hyperlink = u.get('source_url'); c.font = LINK; c.alignment = LEFT
    r += 1
for i, w in enumerate([11,16,40,15,12,20,5,5,6,10,9,11,12,40,52,44,44], 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ---- Sheet 2: Monthly Prices ----
wm = wb.create_sheet('Monthly Prices')
wm['A1'] = 'Nightly price by month — USD (one row per listing)'; wm['A1'].font = Font(name=ARIAL, bold=True, size=14, color='1F4E78')
wm['A2'] = ('One row per unit. Each month = real nightly USD. If a month splits (e.g. "1–19: $98 / 20–31: $100") the price changed mid-month — both real, with the exact days. "blkd" = fully unavailable. Colour: green = low → red = peak.')
wm['A2'].font = NOTE
IDENT = ['wp','Code/Slug','Title','Area','Beds']
hr = 4
for i, c in enumerate(IDENT, 1):
    cell = wm.cell(row=hr, column=i, value=c); cell.font = H; cell.fill = HEAD_FILL; cell.alignment = CEN; cell.border = BORDER
for j, (y, mo) in enumerate(months):
    cell = wm.cell(row=hr, column=len(IDENT)+1+j, value=mlabel(y, mo)); cell.font = H; cell.fill = HEAD_FILL; cell.alignment = CEN; cell.border = BORDER
wm.freeze_panes = wm.cell(row=hr+1, column=len(IDENT)+1).coordinate
r = hr + 1
for u in priced_units:
    wp = u['wp']; cells = [month_cell(wp, y, mo) for (y, mo) in months]
    reps = [c[1] for c in cells if c[1] is not None]
    lo, hi = (min(reps), max(reps)) if reps else (0, 0)
    ident = [wp, u.get('slug'), u.get('title'), u.get('area'), u.get('beds')]
    for i, val in enumerate(ident, 1):
        cell = wm.cell(row=r, column=i, value=val); cell.font = BODY; cell.border = BORDER
        cell.alignment = LEFT if i in (2,3,4) else CEN
    maxlines = 1
    for j, (txt, rep, varying) in enumerate(cells):
        cell = wm.cell(row=r, column=len(IDENT)+1+j, value=txt); cell.border = BORDER
        if rep is not None:
            cell.fill = heat_fill((rep-lo)/(hi-lo)) if hi > lo else NEUTRAL
            if varying:
                cell.font = Font(name=ARIAL, size=8); cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                maxlines = max(maxlines, txt.count('\n') + 1)
            else:
                cell.font = Font(name=ARIAL, size=9); cell.alignment = CEN
        elif txt == 'blkd':
            cell.fill = GREY; cell.font = Font(name=ARIAL, size=8, italic=True, color='909090'); cell.alignment = CEN
        else:
            cell.font = Font(name=ARIAL, size=9); cell.alignment = CEN
    if maxlines > 1: wm.row_dimensions[r].height = 12 * maxlines + 4
    r += 1
for col, w in (('A',9),('B',16),('C',34),('D',14),('E',5)): wm.column_dimensions[col].width = w
for j in range(len(months)): wm.column_dimensions[get_column_letter(len(IDENT)+1+j)].width = 12

# ---- Sheet 3: Price Ranges ----
wr = wb.create_sheet('Price Ranges')
wr['A1'] = 'Nightly price by date range — USD (exact, no estimation)'; wr['A1'].font = Font(name=ARIAL, bold=True, size=14, color='1F4E78')
wr['A2'] = ('Each row = a continuous date range at one flat nightly rate, straight from brassbell.net (no carry-forward, no averaging). Grey "Blocked" rows = nights brassbell shows unavailable — not priced.')
wr['A2'].font = NOTE
rcols = ['wp','Code/Slug','Title','Area','Beds','From','To','Nights','Nightly USD']
hr = 4
for i, c in enumerate(rcols, 1):
    cell = wr.cell(row=hr, column=i, value=c); cell.font = H; cell.fill = HEAD_FILL; cell.alignment = CEN; cell.border = BORDER
wr.freeze_panes = f'A{hr+1}'
r = hr + 1; band = False
for u in priced_units:
    wp = u['wp']; segs = segments(wp); band = not band
    fill = BAND if band else None
    for stt, frm, to in segs:
        nights = (to - frm).days + 1
        row = [wp, u.get('slug'), u.get('title'), u.get('area'), u.get('beds'), frm, to, nights]
        for i, val in enumerate(row, 1):
            cell = wr.cell(row=r, column=i, value=val); cell.border = BORDER
            cell.alignment = LEFT if i in (2,3,4) else CEN
            if fill: cell.fill = fill
        wr.cell(row=r, column=6).number_format = DATE_FMT
        wr.cell(row=r, column=7).number_format = DATE_FMT
        pcell = wr.cell(row=r, column=9)
        if stt[0] == 'P':
            pcell.value = stt[1]; pcell.number_format = USD_FMT; pcell.font = Font(name=ARIAL, size=10, bold=True)
            for i in (1,2,3,4,5,8): wr.cell(row=r, column=i).font = BODY
        else:
            pcell.value = 'Blocked' if stt[0] == 'B' else 'No data'
            for i in range(1, 10):
                cc = wr.cell(row=r, column=i); cc.font = BLK_FONT; cc.fill = GREY
        pcell.alignment = CEN; pcell.border = BORDER
        r += 1
for i, w in enumerate([10,16,38,15,5,15,15,7,13], 1):
    wr.column_dimensions[get_column_letter(i)].width = w

wb.save(OUT)
print(f"built: {len(units_sorted)} units, {len(priced_units)} priced")
print(OUT)
